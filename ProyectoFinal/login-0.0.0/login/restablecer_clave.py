import openpyxl
import re
import os

def obtener_ruta_basededatos():
    # Obtén la ruta del directorio actual del módulo
    ruta_modulo = os.path.dirname(os.path.abspath(__file__))

    # Combina la ruta del directorio del módulo con el nombre de archivo
    ruta_completa = os.path.join(ruta_modulo, "BasedeDatos.xlsx")

    return ruta_completa

def restablecer_clave():
    # Obtener la ruta del archivo BasedeDatos.xlsx
    ruta_completa = obtener_ruta_basededatos()

    # Verificar si el archivo existe
    if not os.path.exists(ruta_completa):
        print("No hay usuarios registrados.")
        return

    # Obtener el nombre de usuario
    usuario = input("Ingrese el nombre de usuario para restablecer la contraseña: ")

    # Abrir el archivo Excel
    wb = openpyxl.load_workbook(ruta_completa)
    ws = wb.active

    # Buscar el usuario en el archivo
    usuario_encontrado = None
    fila_usuario = None
    filas = list(ws.iter_rows(min_row=2, values_only=True))  # Convertir a lista para usar index
    for i, fila in enumerate(filas):
        if fila[0] == usuario:
            usuario_encontrado = fila
            fila_usuario = i + 2  # Sumar 2 porque el índice se basa en 0 y queremos la posición de la fila
            break

    intentos = 0
    while usuario_encontrado and intentos < 2:
        # Solicitar la contraseña actual del usuario
        clave_actual = input("Ingrese la contraseña actual: ")

        # Validar la contraseña actual
        if clave_actual == usuario_encontrado[6]:  # La posición 6 corresponde a la columna de 'CLAVE'
            # Solicitar la nueva contraseña
            nueva_clave = input("Ingrese la nueva contraseña: ")

            # Validar que la nueva contraseña sea alfanumérica y diferente a la contraseña actual
            while not re.match("^(?=.*[A-Za-z])(?=.*\d)[A-Za-z\d]+$", nueva_clave) or nueva_clave == clave_actual:
                if nueva_clave == clave_actual:
                    print("La nueva contraseña no puede ser igual a la contraseña actual. Inténtalo de nuevo.")
                else:
                    print("La contraseña debe ser alfanumérica (contener letras y números). Inténtalo de nuevo.")
                nueva_clave = input("Ingrese la nueva contraseña: ")

            # Actualizar la contraseña en el archivo
            ws.cell(row=fila_usuario, column=7, value=nueva_clave)  # La columna 7 corresponde a la columna de 'CLAVE'

            # Guardar los cambios en el archivo
            wb.save(ruta_completa)

            print(f"Contraseña restablecida con éxito para el usuario {usuario}.")
            break
        else:
            intentos += 1
            print(f"La contraseña actual no es válida. Intento {intentos} de 2.")
            if intentos == 2:
                print("Se alcanzó el número máximo de intentos. No se realizaron cambios.")
                break

if __name__ == "__main__":
    restablecer_clave()
