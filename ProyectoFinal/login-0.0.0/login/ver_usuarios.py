import openpyxl
import os
from tabulate import tabulate

def ocultar_clave_ingreso(valor, encabezado):
    if encabezado in ('CLAVE', 'INGRESO'):
        return '*' * 4
    return valor

def obtener_ruta_basededatos():
    # Obtén la ruta del directorio actual del módulo
    ruta_modulo = os.path.dirname(os.path.abspath(__file__))

    # Combina la ruta del directorio del módulo con el nombre de archivo
    ruta_completa = os.path.join(ruta_modulo, "BasedeDatos.xlsx")

    return ruta_completa

def ver_usuarios():
    # Obtener la ruta del archivo BasedeDatos.xlsx
    ruta_completa = obtener_ruta_basededatos()

    # Verificar si el archivo existe
    if not os.path.exists(ruta_completa):
        print("No hay usuarios registrados.")
        return

    # Abrir el archivo Excel
    wb = openpyxl.load_workbook(ruta_completa)
    ws = wb.active

    # Obtener los encabezados excluyendo las columnas "CLAVE" e "INGRESO"
    encabezados = ws[1]
    encabezados = [encabezado.value for encabezado in encabezados if encabezado.value not in ('CLAVE', 'INGRESO')]

    # Obtener los datos excluyendo las columnas "CLAVE" e "INGRESO"
    datos = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        datos.append([ocultar_clave_ingreso(valor, encabezado) for valor, encabezado in zip(fila, ws[1]) if encabezado.value not in ('CLAVE', 'INGRESO')])

    # Alinear prolijamente las columnas
    tabla = tabulate(datos, headers=encabezados, tablefmt="pipe")

    # Imprimir la tabla
    print(tabla)

if __name__ == "__main__":
    ver_usuarios()
